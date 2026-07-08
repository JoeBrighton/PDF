import streamlit as st
import pdfplumber
import re
import tempfile
import os
import io
import difflib
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Brighton Healthcare Tools", layout="wide")
st.title("Brighton Healthcare Tools")
st.caption("Upload a Clipboard or ShiftKey invoice PDF and an Empion punch report to generate the reconciliation.")
st.info(
    "🔒 **Privacy:** Uploaded files are processed in memory only and never stored. "
    "All data is discarded when you close the browser or upload new files. "
    "Nothing is saved to any server or database.",
    icon=None,
)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def norm(n):
    return n.lower().replace("'", " ").replace("'", " ").replace("-", " ").strip()

def parse_dt(s):
    try:
        return datetime.strptime(s.strip(), "%m/%d/%Y %I:%M %p")
    except:
        return None

def norm_role(r):
    return "CMA" if "CERTIFIED" in r else r

# ─────────────────────────────────────────────────────────────────────────────
# INVOICE PARSER — auto-detects vendor
# ─────────────────────────────────────────────────────────────────────────────

def parse_invoice(pdf_bytes):
    """Detects vendor and routes to the correct parser. Returns (meta, shifts)."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes); tmp_path = tmp.name
    try:
        with pdfplumber.open(tmp_path) as pdf:
            first_page = pdf.pages[0].extract_text() or ""
    finally:
        os.unlink(tmp_path)
    if "ShiftKey" in first_page:
        return parse_shiftkey_invoice(pdf_bytes)
    return parse_clipboard_invoice(pdf_bytes)

def parse_shiftkey_invoice(pdf_bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes); tmp_path = tmp.name
    try:
        with pdfplumber.open(tmp_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    finally:
        os.unlink(tmp_path)

    inv_m    = re.search(r'Invoice\s+(?:Number|#)\s*(\S+)', text)
    fac_m    = re.search(r'ShiftKey,\s+LLC\.\s+Invoice\s*\n(.+)', text)
    period_m = re.search(r'Invoice Period\s+([\w\s,]+?\d{4})\s+-\s+([\w\s,]+?\d{4})', text)
    bal_m    = re.search(r'Balance Due\s+\$([\d,]+\.\d{2})', text)

    def _parse_sk_date(s):
        s = re.sub(r'(\d+)(st|nd|rd|th)', r'\1', s.strip())
        return datetime.strptime(s, "%b %d, %Y")

    def _infer_shift(hour):
        if 6 <= hour < 14:  return "AM"
        if 14 <= hour < 19: return "PM"
        return "NOC"

    meta = {
        'invoice':      inv_m.group(1) if inv_m else '',
        'facility':     fac_m.group(1).strip() if fac_m else '',
        'period_start': _parse_sk_date(period_m.group(1)).strftime("%m/%d/%Y") if period_m else '',
        'period_end':   _parse_sk_date(period_m.group(2)).strftime("%m/%d/%Y") if period_m else '',
        'balance_due':  float(bal_m.group(1).replace(',', '')) if bal_m else 0,
        'vendor':       'ShiftKey',
    }

    MONTHS = r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*'
    ITEM_RE = re.compile(
        MONTHS + r'\s+\d+\w*,\s+\d{4}\s+'        # processed-on date (ignored)
        r'(' + MONTHS + r'\s+\d+\w*,\s+\d{4})\s+' # shift date
        r'(CNA|RN|LPN|LVN)\s+'
        r'([A-Za-z\-\']+,\s+[A-Za-z\-\' ]+?)\s+'  # Last, First
        r'(\d{2}:\d{2})\s+-\s+(\d{2}:\d{2})\s+'   # HH:MM - HH:MM
        r'(\d+)m\s+'                                # break minutes
        r'[\d\s hm]+\s+'                            # duration text (e.g. "7h 30m")
        r'([\d.]+)\s+'                              # hours decimal
        r'\$([\d.]+)\s+'                            # rate
        r'\$([\d,]+\.\d{2})'                        # total
    )

    shifts = []
    for m in ITEM_RE.finditer(text):
        shift_dt = _parse_sk_date(m.group(1))
        date = shift_dt.strftime("%m/%d")
        role = m.group(2)
        parts = m.group(3).split(',', 1)
        emp = f"{parts[1].strip()} {parts[0].strip()}" if len(parts) == 2 else m.group(3).strip()
        t_in_str, t_out_str = m.group(4), m.group(5)
        t_in  = datetime.strptime(f"{shift_dt.strftime('%m/%d/%Y')} {t_in_str}", "%m/%d/%Y %H:%M")
        t_out = datetime.strptime(f"{shift_dt.strftime('%m/%d/%Y')} {t_out_str}", "%m/%d/%Y %H:%M")
        if t_out <= t_in:
            from datetime import timedelta
            t_out += timedelta(days=1)
        inv_hrs = float(m.group(7))  # already break-deducted by ShiftKey
        rate = float(m.group(8))
        shift_label = _infer_shift(t_in.hour)
        shifts.append({
            'date': date, 'role': role, 'emp': emp, 'shift': shift_label,
            'lc': False, 'worked_brk': False, 'inv_hrs': inv_hrs, 'rate': rate,
            'start': t_in, 'end': t_out,
        })
    return meta, shifts

def parse_clipboard_invoice(pdf_bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    try:
        with pdfplumber.open(tmp_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    finally:
        os.unlink(tmp_path)

    inv_num  = re.search(r'INVOICE\s+(\d+(?:-\d+)?)', text)
    facility = re.search(r'Bill To:\s*\n(.+)', text)
    period_s = re.search(r'Period Start Date:\s+(\d{2}/\d{2}/\d{4})', text)
    period_e = re.search(r'Period End Date:\s+(\d{2}/\d{2}/\d{4})', text)
    balance  = re.search(r'Balance Due:\s+\$([\d,]+\.\d{2})', text)

    meta = {
        'invoice':      inv_num.group(1) if inv_num else '',
        'facility':     facility.group(1).strip() if facility else '',
        'period_start': period_s.group(1) if period_s else '',
        'period_end':   period_e.group(1) if period_e else '',
        'balance_due':  float(balance.group(1).replace(',', '')) if balance else 0,
        'vendor':       'Clipboard',
    }

    ROLES = r'(CNA|RN|LPN|LVN|RN/DON|CERTIFIED MEDICATION AIDE)'
    NORMAL_RE = re.compile(
        r'(\d{2}/\d{2}/\d{4}),\s+' + ROLES + r'\s+(.+?),\s+(AM|PM|NOC)\s+([\d.]+)\s+\$([\d.]+)\s+\$([\d,]+\.\d{2})'
    )
    LC_RE = re.compile(
        r'(\d{2}/\d{2}/\d{4}),\s+' + ROLES + r'\s+(.+?),\s+(AM|PM|NOC),\s+\(Late Cancel\)\s+([\d.]+)\s+\$([\d.]+)\s+\$([\d,]+\.\d{2})'
    )
    WB_RE = re.compile(
        r'(\d{2}/\d{2}/\d{4}),\s+' + ROLES + r'\s+(.+?),\s+(AM|PM|NOC)\s+\(Shift has an approved Worked Break Payment Request for\s+([\d.]+)\s+\$([\d.]+)\s+\$([\d,]+\.\d{2})'
    )
    START_RE = re.compile(r'Shift start:\s+(\d{2}/\d{2}/\d{4}\s+\d{1,2}:\d{2}\s+[AP]M)')
    END_RE   = re.compile(r'Shift end:\s+(\d{2}/\d{2}/\d{4}\s+\d{1,2}:\d{2}\s+[AP]M)')

    lines = text.split('\n')
    shifts = []
    for i, line in enumerate(lines):
        lc = wb = False
        m = LC_RE.search(line)
        if m:
            lc = True
        else:
            m = WB_RE.search(line)
            if m:
                wb = True
            else:
                m = NORMAL_RE.search(line)
        if not m:
            continue
        date    = datetime.strptime(m.group(1), "%m/%d/%Y").strftime("%m/%d")
        role    = norm_role(m.group(2))
        emp     = m.group(3).strip()
        shift   = m.group(4)
        inv_hrs = float(m.group(5))
        rate    = float(m.group(6))
        block   = ' '.join(lines[i:i+6])
        ms = START_RE.search(block)
        me = END_RE.search(block)
        shifts.append({
            'date': date, 'role': role, 'emp': emp, 'shift': shift,
            'lc': lc, 'worked_brk': wb, 'inv_hrs': inv_hrs, 'rate': rate,
            'start': parse_dt(ms.group(1)) if ms else None,
            'end':   parse_dt(me.group(1)) if me else None,
        })
    return meta, shifts

# ─────────────────────────────────────────────────────────────────────────────
# EMPION PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_empion(xlsx_bytes):
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = hdr.index('Full Name')
        in_col   = hdr.index('In Time')
        out_col  = hdr.index('Out Time')
        hrs_col  = hdr.index('Total Hours')
        raw = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            fn = row[name_col]; it = row[in_col]; ot = row[out_col]; rh = row[hrs_col]
            if not fn or not it: continue
            p = fn.split(', ', 1)
            fname = f"{p[1]} {p[0]}" if len(p) == 2 else fn
            raw.append({'name': fname, 'date': it.strftime("%m/%d"), 'in': it, 'out': ot, 'raw': rh})
    finally:
        os.unlink(tmp_path)

    combined = defaultdict(list)
    for r in raw:
        combined[(norm(r['name']), r['date'])].append(r)

    emp_idx = {}
    for key, entries in combined.items():
        entries.sort(key=lambda x: x['in'])
        total_raw = sum(e['raw'] for e in entries)
        deduct = int(total_raw // 8) * 0.5
        adj = round(total_raw - deduct, 2)
        emp_idx[key] = {
            'name': entries[0]['name'], 'date': entries[0]['date'],
            'in': entries[0]['in'], 'out': entries[-1]['out'],
            'raw': total_raw, 'deduct': deduct, 'adj': adj,
            'split': len(entries) > 1, 'entries': entries,
        }
    return emp_idx

# ─────────────────────────────────────────────────────────────────────────────
# NAME MATCHING (exact → strip middle name → fuzzy)
# ─────────────────────────────────────────────────────────────────────────────

def find_emp(name, date, emp_idx):
    """Returns (empion_record, match_type, note, confidence_0_to_1)"""
    n = norm(name)
    # 1. Exact
    if (n, date) in emp_idx:
        return emp_idx[(n, date)], 'exact', None, 1.0
    # 2. Drop middle name (keep first + last word only)
    parts = n.split()
    if len(parts) >= 3:
        short = f"{parts[0]} {parts[-1]}"
        if (short, date) in emp_idx:
            matched = emp_idx[(short, date)]
            return matched, 'middle-name-stripped', f"'{name}' matched as '{matched['name']}' (middle name removed)", 0.95
    # 3. Fuzzy match across all names on same date
    date_keys = [k for k in emp_idx if k[1] == date]
    if date_keys:
        emp_names = [k[0] for k in date_keys]
        close = difflib.get_close_matches(n, emp_names, n=1, cutoff=0.75)
        if close:
            match_key = (close[0], date)
            score = difflib.SequenceMatcher(None, n, close[0]).ratio()
            label = 'fuzzy-auto' if score >= 0.85 else 'fuzzy-review'
            matched = emp_idx[match_key]
            return matched, label, f"'{name}' fuzzy-matched to '{matched['name']}' ({score:.0%})", score
    return None, 'no-match', None, 0.0

# ─────────────────────────────────────────────────────────────────────────────
# RECONCILIATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def reconcile(shifts, emp_idx):
    rows = []
    name_notes = []

    # Pre-group invoice lines by (norm_emp, date) so split invoice shifts
    # (same employee, same date, multiple ShiftKey lines) are compared as one total.
    inv_groups = defaultdict(list)
    for s in shifts:
        if s['lc']:
            # Late cancels always stand alone
            rows.append({**s, 'e_in': None, 'e_out': None, 'e_raw': None, 'e_adj': None,
                          'start_diff': None, 'end_diff': None, 'hrs_diff': None,
                          'flag': 'LATE CANCEL', 'notes': '', 'match_type': 'lc', 'confidence': None})
        else:
            inv_groups[(norm(s['emp']), s['date'])].append(s)

    for key, group in inv_groups.items():
        # Sort group by start time so first entry = earliest shift
        group.sort(key=lambda x: x['start'] or datetime.min)
        s = group[0]  # representative shift (earliest)
        date, emp = s['date'], s['emp']
        combined_inv_hrs = round(sum(g['inv_hrs'] for g in group), 2)
        worked_brk = any(g['worked_brk'] for g in group)
        split_inv  = len(group) > 1

        m, match_type, match_note, name_confidence = find_emp(emp, date, emp_idx)
        notes = []
        if split_inv:
            notes.append(f"Split invoice ({len(group)} lines combined: {combined_inv_hrs}h total)")
        if match_note:
            notes.append(match_note)
            name_notes.append(f"{date} {emp}: {match_note}")
        if worked_brk:
            notes.append("Worked break — full gross billed")

        if not m:
            flag = 'MICRO SHIFT' if combined_inv_hrs < 1.0 else 'NO PUNCH'
            # Emit one row per invoice line so punch-for-punch detail is preserved
            for g in group:
                rows.append({**g, 'inv_hrs': combined_inv_hrs if g is s else 0,
                              'e_in': None, 'e_out': None, 'e_raw': None, 'e_adj': None,
                              'start_diff': None, 'end_diff': None, 'hrs_diff': None,
                              'flag': flag, 'notes': '; '.join(notes) if g is s else 'see above',
                              'match_type': 'no-match', 'confidence': 1.0})
            continue

        if m['split']:
            notes.append(f"Split Empion punch ({len(m['entries'])} entries combined)")
        if m['raw'] < 1.0:
            notes.append(f"Micro-punch ({m['raw']}h — likely bad punch)")

        # Compare COMBINED invoice hours vs Empion adjusted hours
        hrs_diff = round(combined_inv_hrs - m['adj'], 2)
        if abs(hrs_diff) < 0.05:    flag = 'MATCH'
        elif hrs_diff > 0:          flag = f'OVERBILLED +{hrs_diff:.2f}h'
        else:                       flag = f'UNDERBILLED {hrs_diff:.2f}h'
        if worked_brk:              flag = 'WORKED BREAK'
        if m['raw'] < 1.0:         flag = 'BAD PUNCH'
        if match_type == 'fuzzy-review': flag = f'REVIEW NAME MATCH | {flag}'
        confidence = round(name_confidence, 2)

        # Earliest inv start / latest inv end across all lines in group
        inv_start = min((g['start'] for g in group if g['start']), default=None)
        inv_end   = max((g['end']   for g in group if g['end']),   default=None)
        start_diff = round((m['in']  - inv_start).total_seconds() / 60, 1) if inv_start else None
        end_diff   = round((m['out'] - inv_end  ).total_seconds() / 60, 1) if inv_end   else None

        # Emit one summary row (combined inv_hrs) for the group
        rows.append({**s, 'inv_hrs': combined_inv_hrs, 'start': inv_start, 'end': inv_end,
                     'e_in': m['in'], 'e_out': m['out'], 'e_raw': m['raw'], 'e_adj': m['adj'],
                     'start_diff': start_diff, 'end_diff': end_diff,
                     'hrs_diff': hrs_diff, 'flag': flag, 'notes': '; '.join(notes),
                     'match_type': match_type, 'confidence': confidence})

    # Sort output by date then employee name for readability
    rows.sort(key=lambda r: (r['date'], r['emp']))
    return rows, name_notes

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def make_styles():
    return {
        'HDR_F': PatternFill("solid", start_color="1F4E79"),
        'HFONT': Font(name="Arial", bold=True, color="FFFFFF", size=9),
        'RED_F': PatternFill("solid", start_color="FFD7D7"),
        'YLW_F': PatternFill("solid", start_color="FFE699"),
        'ORG_F': PatternFill("solid", start_color="FCE4D6"),
        'GRN_F': PatternFill("solid", start_color="E2EFDA"),
        'GRY_F': PatternFill("solid", start_color="F2F2F2"),
        'UND_F': PatternFill("solid", start_color="FFF2CC"),
        'BLU_F': PatternFill("solid", start_color="DDEBF7"),
        'NF':  Font(name="Arial", size=9),
        'BF':  Font(name="Arial", bold=True, size=9),
        'RF':  Font(name="Arial", bold=True, size=9, color="C00000"),
        'BLF': Font(name="Arial", bold=True, size=9, color="1F4E79"),
        'CTR': Alignment(horizontal="center", vertical="center"),
        'LFT': Alignment(horizontal="left",   vertical="center"),
        'WRAP': Alignment(horizontal="left",  vertical="center", wrap_text=True),
        'TFmt': "M/D H:MM AM/PM",
        'RFILLS': {
            'RN':  PatternFill("solid", start_color="E2EFDA"),
            'LPN': PatternFill("solid", start_color="DDEBF7"),
            'LVN': PatternFill("solid", start_color="DDEBF7"),
            'CNA': PatternFill("solid", start_color="FCE4D6"),
            'CMA': PatternFill("solid", start_color="FFF2CC"),
        },
    }

def build_hours_excel(meta, shifts):
    S = make_styles()
    ROLE_ORDER = {'RN': 0, 'LPN': 1, 'LVN': 1, 'CMA': 2, 'CNA': 3}
    RATES = {'RN': 58, 'LPN': 45, 'LVN': 45, 'CMA': 28.5, 'CNA': 27}

    # Collect dates and employees
    all_dates = sorted(set(s['date'] for s in shifts))
    emp_role = {}
    for s in shifts:
        if s['emp'] not in emp_role:
            emp_role[s['emp']] = s['role']
    employees = sorted(emp_role, key=lambda n: (ROLE_ORDER.get(emp_role[n], 9), n))
    grid = defaultdict(lambda: defaultdict(list))
    for s in shifts:
        grid[s['emp']][s['date']].append(s['inv_hrs'])

    wb = Workbook(); ws = wb.active; ws.title = "Hours by Day"
    ncols = 2 + len(all_dates) + 2
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = f"Clipboard Invoice {meta['invoice']} — {meta['facility']} — {meta['period_start']}–{meta['period_end']}"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws.row_dimensions[1].height = 22

    from calendar import day_abbr
    def day_label(d):
        try: dt = datetime.strptime(d + "/2026", "%m/%d/%Y"); return f"{dt.strftime('%a')} {d}"
        except: return d

    hdrs = ["Employee", "Role"] + [day_label(d) for d in all_dates] + ["Total Hrs", "Total $"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = S['HFONT']; c.fill = S['HDR_F']; c.alignment = S['CTR']
    ws.row_dimensions[2].height = 18

    row = 3
    for emp in employees:
        role = emp_role[emp]
        rf = S['RFILLS'].get(role, PatternFill())
        ws.cell(row=row, column=1, value=emp).font = S['NF']
        ws.cell(row=row, column=1).alignment = S['LFT']
        ws.cell(row=row, column=1).fill = rf
        ws.cell(row=row, column=2, value=role).alignment = S['CTR']
        ws.cell(row=row, column=2).fill = rf
        ws.cell(row=row, column=2).font = Font(name="Arial", size=9, bold=(role == "RN"))
        th = 0; ta = 0
        for di, d in enumerate(all_dates):
            hrs_list = grid[emp].get(d, [])
            if hrs_list:
                hrs = sum(hrs_list); th += hrs; ta += hrs * RATES.get(role, 27)
                c = ws.cell(row=row, column=3+di, value=round(hrs, 2))
                c.fill = rf; c.alignment = S['CTR']; c.number_format = "0.00"; c.font = S['NF']
            else:
                ws.cell(row=row, column=3+di, value="").fill = rf
        bl = PatternFill("solid", start_color="D6E4F0")
        tc = ws.cell(row=row, column=3+len(all_dates), value=round(th, 2))
        tc.font = S['BF']; tc.alignment = S['CTR']; tc.fill = bl; tc.number_format = "0.00"
        tm = ws.cell(row=row, column=4+len(all_dates), value=round(ta, 2))
        tm.font = S['BF']; tm.alignment = Alignment(horizontal="right", vertical="center")
        tm.fill = bl; tm.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 15; row += 1

    tr = row
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    for ci in range(1, ncols+1):
        ws.cell(row=tr, column=ci).fill = S['HDR_F']
    for ci in range(3, ncols+1):
        col = get_column_letter(ci)
        c = ws.cell(row=tr, column=ci, value=f"=SUM({col}3:{col}{tr-1})")
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = S['HDR_F']; c.alignment = S['CTR']
        c.number_format = '"$"#,##0.00' if ci == ncols else "0.00"
    ws.row_dimensions[tr].height = 18

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 6
    for ci in range(3, 3+len(all_dates)): ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.column_dimensions[get_column_letter(3+len(all_dates))].width = 11
    ws.column_dimensions[get_column_letter(4+len(all_dates))].width = 13

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def build_flags_excel(meta, recon_rows):
    S = make_styles()

    def row_fill(flag):
        if 'NO PUNCH' in flag:    return S['RED_F']
        if 'BAD PUNCH' in flag:   return S['YLW_F']
        if 'WORKED BREAK' in flag: return S['YLW_F']
        if 'OVERBILLED' in flag:  return S['ORG_F']
        if 'UNDERBILLED' in flag: return S['UND_F']
        if 'REVIEW' in flag:      return S['BLU_F']
        if flag == 'MATCH':       return S['GRN_F']
        if flag == 'LATE CANCEL': return S['GRY_F']
        return S['GRN_F']

    # Pre-compute summary stats
    total_items    = len(recon_rows)
    n_match  = sum(1 for r in recon_rows if r['flag'] == 'MATCH')
    n_lc     = sum(1 for r in recon_rows if r['flag'] == 'LATE CANCEL')
    n_np     = sum(1 for r in recon_rows if 'NO PUNCH' in r['flag'])
    n_over   = sum(1 for r in recon_rows if 'OVERBILLED' in r['flag'])
    n_under  = sum(1 for r in recon_rows if 'UNDERBILLED' in r['flag'])
    n_other  = total_items - n_match - n_lc - n_np - n_over - n_under
    total_inv_hrs  = round(sum(r['inv_hrs'] for r in recon_rows if not r['lc']), 2)
    total_emp_hrs  = round(sum(r['e_adj'] for r in recon_rows if r['e_adj'] is not None), 2)
    punch_disc_hrs = round(sum(r['hrs_diff'] for r in recon_rows if r.get('hrs_diff') is not None), 2)
    missing_hrs    = round(sum(r['inv_hrs'] for r in recon_rows if r['match_type'] == 'no-match'), 2)

    wb = Workbook()

    # ── Sheet 1: Flags ────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Flags"

    # Title
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"Invoice {meta['invoice']} — {meta['facility']} — {meta.get('vendor','Clipboard')}"
    ws1["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws1.row_dimensions[1].height = 22

    # Period / balance
    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"Period: {meta['period_start']} – {meta['period_end']}   |   Balance Due: ${meta['balance_due']:,.2f}"
    ws1["A2"].font = Font(name="Arial", size=9, color="595959")
    ws1.row_dimensions[2].height = 14

    # Summary stats header row
    SUMHDR_F = PatternFill("solid", start_color="2E75B6")
    SUMVAL_F = PatternFill("solid", start_color="D6E4F0")
    stat_labels = ["Match", "No Punch", "Overbilled", "Underbilled", "Other", "Late Cancel",
                   "Inv Hours", "Emp Hours", "Punch Disc.", "Missing Hrs"]
    stat_values = [n_match, n_np, n_over, n_under, n_other, n_lc,
                   f"{total_inv_hrs:.2f}h", f"{total_emp_hrs:.2f}h",
                   f"{punch_disc_hrs:+.2f}h", f"{missing_hrs:.2f}h"]
    for ci, lbl in enumerate(stat_labels, 1):
        c = ws1.cell(row=3, column=ci, value=lbl)
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill = SUMHDR_F; c.alignment = Alignment(horizontal="center", vertical="center")
    for ci, val in enumerate(stat_values, 1):
        c = ws1.cell(row=4, column=ci, value=val)
        c.font = Font(name="Arial", bold=True, size=9)
        c.fill = SUMVAL_F; c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 14
    ws1.row_dimensions[4].height = 16

    # Blank spacer
    ws1.row_dimensions[5].height = 6

    # Column headers (shifted down by 4 rows)
    for ci, h in enumerate(["Date","Employee","Role","Issue","Invoice In","Invoice Out","Empion In","Empion Out","Notes"], 1):
        c = ws1.cell(row=6, column=ci, value=h)
        c.font = S['HFONT']; c.fill = S['HDR_F']; c.alignment = S['CTR']
    ws1.row_dimensions[6].height = 16

    ri = 7
    for r in recon_rows:
        if r['flag'] in ('MATCH', 'LATE CANCEL'): continue
        fill = row_fill(r['flag'])
        def s1(col, val, fmt=None, fnt=None, aln=None):
            c = ws1.cell(row=ri, column=col, value=val)
            c.fill = fill; c.alignment = aln or S['CTR']; c.font = fnt or S['NF']
            if val and fmt: c.number_format = fmt
        s1(1, r['date'], fnt=S['BF'])
        s1(2, r['emp'], aln=S['LFT'], fnt=S['BF'])
        s1(3, r['role'])
        fc = ws1.cell(row=ri, column=4, value=r['flag'])
        fc.fill = fill; fc.alignment = S['CTR']
        fc.font = S['RF'] if ('OVER' in r['flag'] or 'NO PUNCH' in r['flag']) else S['BF']
        s1(5, r['start'], S['TFmt']); s1(6, r['end'], S['TFmt'])
        s1(7, r['e_in'], S['TFmt']); s1(8, r['e_out'], S['TFmt'])
        for col in [5, 6, 7, 8]:
            c = ws1.cell(row=ri, column=col)
            if c.value is None: c.value = "—"
        s1(9, r['notes'], aln=S['WRAP'])
        ws1.row_dimensions[ri].height = 18; ri += 1

    ws1.column_dimensions["A"].width = 7;  ws1.column_dimensions["B"].width = 26
    ws1.column_dimensions["C"].width = 5;  ws1.column_dimensions["D"].width = 24
    for col in ["E","F","G","H"]: ws1.column_dimensions[col].width = 18
    ws1.column_dimensions["I"].width = 48
    ws1.freeze_panes = "A7"

    # ── Sheet 2: Punch-for-Punch ──────────────────────────────────────────
    ws2 = wb.create_sheet("Punch-for-Punch")
    ws2.merge_cells("A1:P1")
    ws2["A1"] = f"Invoice {meta['invoice']} — {meta['facility']} — Full Punch-for-Punch vs Empion"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    ws2.row_dimensions[1].height = 20

    for ci, h in enumerate(["Date","Employee","Role","Shift","Inv Start","Inv End","Inv Gross","Inv Billed",
                             "Emp In","Emp Out","Emp Raw","Emp Adj","Start Δ(min)","End Δ(min)","Hrs Diff","Flag / Notes"], 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = S['HFONT']; c.fill = S['HDR_F']
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[2].height = 30

    for ri2, r in enumerate(recon_rows, 3):
        fill = row_fill(r['flag'])
        i_gross = round((r['end'] - r['start']).total_seconds() / 3600, 2) if r['start'] and r['end'] else None
        def cv2(col, val, fmt=None, fnt=None, aln=None):
            c = ws2.cell(row=ri2, column=col, value=val)
            c.fill = fill; c.alignment = aln or S['CTR']; c.font = fnt or S['NF']
            if fmt and val is not None: c.number_format = fmt
        conf = r.get('confidence')
        conf_str = f" | Confidence: {conf:.0%}" if conf is not None else ""
        note_val = r['flag'] + conf_str + (' | ' + r['notes'] if r['notes'] else '')
        cv2(1, r['date']); cv2(2, r['emp'], aln=S['LFT']); cv2(3, r['role']); cv2(4, r['shift'])
        cv2(5, r['start'], S['TFmt']); cv2(6, r['end'], S['TFmt'])
        cv2(7, i_gross, "0.00"); cv2(8, r['inv_hrs'], "0.00", fnt=S['BF'])
        cv2(9, r['e_in'], S['TFmt']); cv2(10, r['e_out'], S['TFmt'])
        cv2(11, r['e_raw'], "0.00"); cv2(12, r['e_adj'], "0.00", fnt=S['BF'])
        sd = ws2.cell(row=ri2, column=13, value=r['start_diff'])
        sd.fill = fill; sd.alignment = S['CTR']; sd.number_format = "+0.0;-0.0;0"
        sd.font = S['RF'] if (r['start_diff'] and abs(r['start_diff']) >= 10) else S['NF']
        ed = ws2.cell(row=ri2, column=14, value=r['end_diff'])
        ed.fill = fill; ed.alignment = S['CTR']; ed.number_format = "+0.0;-0.0;0"
        ed.font = S['RF'] if (r['end_diff'] and abs(r['end_diff']) >= 10) else S['NF']
        hc = ws2.cell(row=ri2, column=15, value=r['hrs_diff'])
        hc.fill = fill; hc.alignment = S['CTR']; hc.number_format = "+0.00;-0.00;0.00"
        if r['hrs_diff'] and r.get('e_raw') and r['e_raw'] >= 1 and not r['worked_brk']:
            hc.font = S['RF'] if r['hrs_diff'] > 0.05 else (S['BLF'] if r['hrs_diff'] < -0.05 else S['NF'])
        else:
            hc.font = S['NF']
        fc = ws2.cell(row=ri2, column=16, value=note_val)
        fc.fill = fill; fc.alignment = S['WRAP']; fc.font = S['NF']
        ws2.row_dimensions[ri2].height = 14

    for i, w in enumerate([7,26,5,6,17,17,9,9,17,17,9,9,12,12,10,42], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────────────────
# AP PAYMENT RECONCILIATION
# ─────────────────────────────────────────────────────────────────────────────

def parse_intact_excel(xlsx_bytes):
    """Parse Intact GL register export — all transaction types."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Find header row (must contain 'Amount'; 'Check No' optional)
    hdr_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v).strip() if v is not None else '' for v in row]
        if 'Amount' in vals:
            hdr_row = i
            headers = vals
            break
    if not hdr_row:
        raise ValueError("Could not find header row with 'Amount'")

    def col(name):
        return next((i for i, h in enumerate(headers) if name.lower() in h.lower()), None)

    ck_col   = col('Check No')
    amt_col  = headers.index('Amount')
    vnd_col  = col('Vendor')
    dt_col   = col('GL date') if col('GL date') is not None else col('GL Date')
    type_col = col('Checks and debits') if col('Checks and debits') is not None else col('transaction type')
    sub_col  = col('AP payment') if col('AP payment') is not None else col('subtype')

    rows = []
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        amt = row[amt_col] if amt_col is not None else None
        if amt is None:
            continue
        try:
            amt = float(amt)
        except (ValueError, TypeError):
            continue
        if amt == 0:
            continue

        ck = row[ck_col] if ck_col is not None else None
        if ck is not None:
            try:
                ck = str(int(float(ck)))
            except (ValueError, TypeError):
                ck = str(ck).strip() or None

        vendor = str(row[vnd_col]).strip() if vnd_col is not None and row[vnd_col] else ''
        gl_date = ''
        if dt_col is not None and row[dt_col]:
            v = row[dt_col]
            gl_date = v.strftime('%m/%d/%Y') if hasattr(v, 'strftime') else str(v)

        txn_type = str(row[type_col]).strip() if type_col is not None and row[type_col] else ''
        txn_sub  = str(row[sub_col]).strip()  if sub_col  is not None and row[sub_col]  else ''

        rows.append({
            'check_no': ck,
            'amount':   amt,
            'vendor':   vendor,
            'gl_date':  gl_date,
            'txn_type': txn_type,
            'txn_sub':  txn_sub,
        })
    return rows

def parse_bank_statement(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    transactions = {}

    LINE_RE  = re.compile(r'^(\d{2}/\d{2})\s+(.+?)\s+([\d,]+\.\d{2})$')
    CK_V_RE  = re.compile(r'Ck#\s+V(\d+)')

    # ── WITHDRAWALS section (captures every debit line generically) ──────────
    in_wd = False
    wire_idx = 1
    for line in text.split('\n'):
        s = line.strip()
        if re.match(r'^WITHDRAWALS', s) and 'Total' not in s:
            in_wd = True; continue
        if 'Total Withdrawals' in s or re.match(r'^FEES', s):
            in_wd = False
        if not in_wd:
            continue
        m = LINE_RE.match(s)
        if not m:
            continue
        date, desc, amt_str = m.group(1), m.group(2), m.group(3)
        amount = float(amt_str.replace(',', ''))

        ck_m = CK_V_RE.search(desc)
        if ck_m:
            # Intact ACH check (Ck# V###)
            ck_no = ck_m.group(1)
            # extract vendor between "Ck# Vxxx Chrr" and "Whigham"
            vend_m = re.search(r'Ck#\s+V\d+\s+Chrr\s+(.+?)\s+(?:Whigham|ACH\.Live)', desc)
            vendor_raw = vend_m.group(1).strip() if vend_m else desc
            transactions[ck_no] = {
                'date': date, 'check_no': ck_no,
                'vendor_raw': vendor_raw, 'amount': amount, 'type': 'ACH Check',
            }
        elif 'Wire Transfer' in desc:
            key = f"WIRE_{wire_idx}"; wire_idx += 1
            vendor_raw = desc.replace('Wire Transfer', '').strip()
            transactions[key] = {
                'date': date, 'check_no': None,
                'vendor_raw': vendor_raw, 'amount': amount, 'type': 'Wire',
            }
        else:
            # General ACH / debit / merchant charge
            # Extract numeric reference ID from description (e.g. "14003100000392")
            ref_m = re.search(r'(?<!\d)(\d{8,})(?!\d)', desc)
            ref_id = ref_m.group(1) if ref_m else None
            key = f"ACH_{date}_{len(transactions)}"
            transactions[key] = {
                'date': date, 'check_no': None,
                'vendor_raw': desc, 'amount': amount, 'type': 'ACH/Debit',
                'ref_id': ref_id,
            }

    # ── CHECKS section (paper checks — date + check# + amount columns) ──────
    in_checks = False
    CHK_SEC_RE = re.compile(r'(\d{2}/\d{2})\s+(\d{3,})\s+\*?\s*([\d,]+\.\d{2})')
    for line in text.split('\n'):
        s = line.strip()
        if re.match(r'^CHECKS', s) and 'CONTINUED' not in s and 'Total' not in s:
            in_checks = True; continue
        if 'Total Checks' in s:
            in_checks = False
        if in_checks:
            for m in CHK_SEC_RE.finditer(s):
                ck_no = m.group(2)
                transactions[ck_no] = {
                    'date': m.group(1), 'check_no': ck_no,
                    'vendor_raw': '', 'amount': float(m.group(3).replace(',', '')),
                    'type': 'Check',
                }
    return transactions

def match_intact_to_bank(intact_rows, bank_txns):
    """
    Match Intact rows to bank transactions — NO date used (accrual accounting).
    Priority:
      1. Check number exact match (Intact check_no == bank check_no)
      2. Amount exact match (within $0.02) + vendor fuzzy match
      3. Amount exact match alone (last resort)
    Each bank transaction can only be matched once.
    """
    used_bank = set()  # id() of already-matched bank txns
    matched = []

    # Pass 1: check number
    for row in intact_rows:
        ck = str(row.get('check_no', '')).strip()
        amt = row.get('amount')
        if ck and ck in bank_txns:
            b = bank_txns[ck]
            if id(b) not in used_bank and (amt is None or abs(b['amount'] - amt) < 0.02):
                matched.append({'intact_row': row, 'bank': b, 'match_type': 'check_no'})
                used_bank.add(id(b))

    matched_intact = {id(m['intact_row']) for m in matched}

    # Pass 2: amount + vendor fuzzy
    for row in intact_rows:
        if id(row) in matched_intact:
            continue
        amt = row.get('amount')
        vendor = norm(row.get('vendor', ''))
        if not amt:
            continue
        best_score, best_b = 0, None
        for b in bank_txns.values():
            if id(b) in used_bank:
                continue
            if abs(b['amount'] - amt) < 0.02:
                score = difflib.SequenceMatcher(None, vendor, norm(b.get('vendor_raw', ''))).ratio()
                if score > best_score:
                    best_score, best_b = score, b
        if best_b and best_score >= 0.4:
            matched.append({'intact_row': row, 'bank': best_b, 'match_type': f'amount+vendor ({best_score:.0%})'})
            used_bank.add(id(best_b))
            matched_intact.add(id(row))

    # Pass 3: amount only
    for row in intact_rows:
        if id(row) in matched_intact:
            continue
        amt = row.get('amount')
        if not amt:
            continue
        for b in bank_txns.values():
            if id(b) in used_bank:
                continue
            if abs(b['amount'] - amt) < 0.02:
                matched.append({'intact_row': row, 'bank': b, 'match_type': 'amount only'})
                used_bank.add(id(b))
                matched_intact.add(id(row))
                break

    # Pass 4: split bank transactions — same reference ID, sum matches Intact amount
    # Group unmatched bank txns by ref_id
    from collections import defaultdict
    ref_groups = defaultdict(list)
    for b in bank_txns.values():
        if id(b) not in used_bank and b.get('ref_id'):
            ref_groups[b['ref_id']].append(b)
    # Only keep groups with 2+ transactions
    ref_groups = {k: v for k, v in ref_groups.items() if len(v) >= 2}

    for row in intact_rows:
        if id(row) in matched_intact:
            continue
        amt = row.get('amount')
        if not amt:
            continue
        for ref_id, group in ref_groups.items():
            group_sum = sum(b['amount'] for b in group)
            if abs(group_sum - amt) < 0.02:
                # Match — use the largest transaction as the primary bank entry
                primary = max(group, key=lambda b: b['amount'])
                split_desc = ' + '.join(b['vendor_raw'][:20] for b in sorted(group, key=lambda b: b['amount'], reverse=True))
                matched.append({
                    'intact_row': row,
                    'bank': primary,
                    'match_type': f'split ({len(group)} bank txns, ref {ref_id})',
                    'split_group': group,
                })
                for b in group:
                    used_bank.add(id(b))
                matched_intact.add(id(row))
                break

    return matched

def ocr_intact_screenshot(img_bytes):
    """OCR the Intact screenshot, return list of row dicts with bounding boxes."""
    try:
        import pytesseract
        from PIL import Image
        import subprocess
        # verify tesseract binary is accessible
        result = subprocess.run(['tesseract', '--version'], capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(f"Tesseract binary not found: {result.stderr}")
        pytesseract.pytesseract.tesseract_cmd = 'tesseract'
        img = Image.open(io.BytesIO(img_bytes))
        data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DATAFRAME)
        import pandas as pd
        data = data[data.conf > 30].dropna(subset=['text'])
        data = data[data['text'].str.strip() != '']

        # Group words into rows        # Group words into rows by Y coordinate (cluster within 8px)
        data = data.sort_values('top')
        rows = []
        current_row_top = None
        current_row_words = []

        for _, word in data.iterrows():
            if current_row_top is None or abs(word['top'] - current_row_top) > 8:
                if current_row_words:
                    rows.append(current_row_words)
                current_row_words = [word]
                current_row_top = word['top']
            else:
                current_row_words.append(word)

        if current_row_words:
            rows.append(current_row_words)

        # Extract structured data from each row
        result = []
        AMT_RE = re.compile(r'^[\d,]+\.\d{2}$')
        CK_RE  = re.compile(r'^\d{3,6}$')

        for i, words in enumerate(rows):
            texts = [w['text'] for w in words]
            line  = ' '.join(texts)
            top   = min(w['top'] for w in words)
            bot   = max(w['top'] + w['height'] for w in words)

            # Rejoin comma-split amounts (OCR may split "2,382.30" into "2,382" + ".30")
            rejoined = re.findall(r'[\d,]+\.\d{2}', ' '.join(texts))
            amounts = [float(t.replace(',', '')) for t in rejoined if AMT_RE.match(t)]
            ck_nums = [t for t in texts if CK_RE.match(t)]

            result.append({
                'row_index': i,
                'line': line,
                'top': top,
                'bottom': bot,
                'check_no': ck_nums[0] if ck_nums else None,
                'amount': amounts[0] if amounts else None,
            })
        return img, result

    except Exception as e:
        return None, [{'_error': str(e)}]


def annotate_screenshot(img, matched_row_tops, img_width):
    """Draw semi-transparent green highlights over matched rows."""
    from PIL import Image, ImageDraw
    overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    for top, bot in matched_row_tops:
        draw.rectangle([(0, top - 2), (img_width, bot + 2)],
                       fill=(0, 200, 80, 80))
    base = img.convert('RGBA')
    combined = Image.alpha_composite(base, overlay)
    buf = io.BytesIO()
    combined.convert('RGB').save(buf, format='PNG')
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

tab1, tab2 = st.tabs(["Staffing Invoice Recon", "AP Payment Recon"])

with tab2:
    st.subheader("AP Payment Reconciliation")
    st.caption("Upload the Intact uncleared checks Excel and your bank statement PDF.")
    st.info("Files are processed in memory only and never stored.", icon=None)

    ap1, ap2 = st.columns(2)
    with ap1:
        intact_file = st.file_uploader("Intact Uncleared Checks (Excel)", type=["xlsx", "xls"], key="intact")
    with ap2:
        bank_file = st.file_uploader("Bank Statement (PDF)", type="pdf", key="bank")

    if intact_file and bank_file:
        if st.button("Run AP Reconciliation", type="primary", use_container_width=True):
            with st.spinner("Parsing Intact export..."):
                intact_rows = parse_intact_excel(intact_file.read())
            with st.spinner("Parsing bank statement..."):
                bank_txns = parse_bank_statement(bank_file.read())
            with st.spinner("Matching..."):
                matches = match_intact_to_bank(intact_rows, bank_txns)

            matched_by_intact_id = {id(m['intact_row']): m for m in matches}

            st.divider()
            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Intact items",        len(intact_rows))
            mc2.metric("Bank transactions",   len(bank_txns))
            mc3.metric("Cleared",             len(matches))
            mc4.metric("Still pending",       len(intact_rows) - len(matches))

            import pandas as pd
            all_rows = []
            for r in intact_rows:
                m = matched_by_intact_id.get(id(r))
                bank = m['bank'] if m else None
                all_rows.append({
                    'Status':      'CLEARED' if m else 'PENDING',
                    'Match Type':  m['match_type'] if m else '',
                    'Check #':     r['check_no'] or '--',
                    'GL Date':     r['gl_date'],
                    'Txn Type':    r.get('txn_sub') or r.get('txn_type') or '',
                    'Vendor':      r['vendor'],
                    'Amount':      r['amount'],
                    'Bank Date':   bank['date'] if bank else '',
                    'Bank Type':   bank['type'] if bank else '',
                })

            df = pd.DataFrame(all_rows)
            st.dataframe(df.style.apply(
                lambda row: ['background-color: #d4edda' if row['Status'] == 'CLEARED'
                             else 'background-color: #f8d7da' for _ in row], axis=1
            ), use_container_width=True, hide_index=True)

            # Pending vendors summary
            pending = [r for r in all_rows if r['Status'] == 'PENDING']
            if pending:
                from collections import defaultdict
                vendor_totals = defaultdict(float)
                vendor_checks = defaultdict(int)
                for r in pending:
                    vendor_totals[r['Vendor']] += r['Amount']
                    vendor_checks[r['Vendor']] += 1
                vendor_summary = sorted(vendor_totals.items(), key=lambda x: x[1], reverse=True)
                st.subheader(f"{len(pending)} Pending Items — {len(vendor_summary)} Vendor(s)")
                pending_df = pd.DataFrame([
                    {'Vendor': v, 'Checks': vendor_checks[v], 'Outstanding ($)': f"${amt:,.2f}"}
                    for v, amt in vendor_summary
                ])
                st.dataframe(pending_df, use_container_width=True, hide_index=True)
            else:
                st.success("All items cleared!")

            # Bank transactions not in Intact
            matched_bank_keys = {m['bank']['check_no'] or f"WIRE_{list(bank_txns.keys()).index(next(k for k,v in bank_txns.items() if v is m['bank']))}" for m in matches}
            matched_bank_txns = {id(m['bank']) for m in matches}
            unmatched_bank = [v for v in bank_txns.values() if id(v) not in matched_bank_txns]
            if unmatched_bank:
                st.divider()
                st.subheader(f"{len(unmatched_bank)} Bank Transaction(s) Not in Intact")
                st.caption("These cleared the bank but have no matching Intact uncleared check — may already be posted or entered outside Intact.")
                ub_df = pd.DataFrame([{
                    'Bank Date':  t['date'],
                    'Check #':    t['check_no'] or '--',
                    'Vendor':     t['vendor_raw'] or '--',
                    'Amount':     f"${t['amount']:,.2f}",
                    'Type':       t['type'],
                } for t in sorted(unmatched_bank, key=lambda x: x['amount'], reverse=True)])
                st.dataframe(ub_df, use_container_width=True, hide_index=True)

            # Download Excel
            out = io.BytesIO()
            from openpyxl import Workbook as WB2
            wb2 = WB2(); ws2 = wb2.active; ws2.title = "AP Reconciliation"
            GREEN = PatternFill("solid", start_color="C6EFCE")
            RED_F = PatternFill("solid", start_color="FFC7CE")
            hdrs = list(all_rows[0].keys())
            for ci, h in enumerate(hdrs, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                c.fill = PatternFill("solid", start_color="1F4E79")
            for ri, row in enumerate(all_rows, 2):
                fill = GREEN if row['Status'] == 'CLEARED' else RED_F
                for ci, key in enumerate(hdrs, 1):
                    c = ws2.cell(row=ri, column=ci, value=row[key])
                    c.fill = fill
                    c.font = Font(name="Arial", size=9)
            for ci in [1,2,3]: ws2.column_dimensions[get_column_letter(ci)].width = 12
            ws2.column_dimensions["D"].width = 32
            for ci in [5,6,7]: ws2.column_dimensions[get_column_letter(ci)].width = 14
            # Sheet 2: Not in Intact checklist
            ws3 = wb2.create_sheet("Not in Intact")
            ni_hdrs = ['', 'Bank Date', 'Type', 'Check #', 'Vendor', 'Amount', 'Notes']
            for ci, h in enumerate(ni_hdrs, 1):
                c = ws3.cell(row=1, column=ci, value=h)
                c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                c.fill = PatternFill("solid", start_color="1F4E79")
            YELLOW = PatternFill("solid", start_color="FFEB9C")
            unmatched_sorted = sorted(unmatched_bank, key=lambda x: x['amount'], reverse=True)
            for ri, t in enumerate(unmatched_sorted, 2):
                ws3.cell(row=ri, column=1, value='☐').font = Font(name="Arial", size=11)
                ws3.cell(row=ri, column=2, value=t['date']).font = Font(name="Arial", size=9)
                ws3.cell(row=ri, column=3, value=t['type']).font = Font(name="Arial", size=9)
                ws3.cell(row=ri, column=4, value=t['check_no'] or '--').font = Font(name="Arial", size=9)
                ws3.cell(row=ri, column=5, value=t['vendor_raw'] or '--').font = Font(name="Arial", size=9)
                amt_cell = ws3.cell(row=ri, column=6, value=t['amount'])
                amt_cell.font = Font(name="Arial", size=9)
                amt_cell.number_format = '"$"#,##0.00'
                notes_cell = ws3.cell(row=ri, column=7, value='')
                notes_cell.fill = YELLOW
                notes_cell.font = Font(name="Arial", size=9)
            ws3.column_dimensions["A"].width = 4
            ws3.column_dimensions["B"].width = 10
            ws3.column_dimensions["C"].width = 12
            ws3.column_dimensions["D"].width = 10
            ws3.column_dimensions["E"].width = 40
            ws3.column_dimensions["F"].width = 14
            ws3.column_dimensions["G"].width = 30
            ws3.freeze_panes = "A2"

            wb2.save(out); out.seek(0)
            st.download_button(
                "Download Reconciliation (Excel)",
                data=out.read(),
                file_name="AP_Reconciliation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.info("Upload both files above to get started.")

with tab1:
    st.caption("Upload a Clipboard or ShiftKey invoice PDF and an Empion punch report to generate the reconciliation.")
    st.info(
        "Privacy: Uploaded files are processed in memory only and never stored. "
        "All data is discarded when you close the browser or upload new files. "
        "Nothing is saved to any server or database.",
        icon=None,
    )

    col1, col2 = st.columns(2)
    with col1:
        pdf_file  = st.file_uploader("Invoice PDF (Clipboard or ShiftKey)", type="pdf")
    with col2:
        xlsx_file = st.file_uploader("Empion Punch Report (Excel)", type=["xlsx", "xls"])

    if pdf_file and xlsx_file:
        if st.button("Run Reconciliation", type="primary", use_container_width=True):
            with st.spinner("Detecting vendor and parsing invoice..."):
                meta, shifts = parse_invoice(pdf_file.read())
                vendor = meta.get('vendor', 'Clipboard')
            with st.spinner("Parsing Empion punches..."):
                emp_idx = parse_empion(xlsx_file.read())
            with st.spinner("Reconciling..."):
                recon_rows, name_notes = reconcile(shifts, emp_idx)

            st.divider()
            st.subheader(f"{vendor} Invoice {meta['invoice']} -- {meta['facility']}")
            st.caption(f"Period: {meta['period_start']} - {meta['period_end']}  |  Balance Due: ${meta['balance_due']:,.2f}")

            total_items = len(recon_rows)
            n_match  = sum(1 for r in recon_rows if r['flag'] == 'MATCH')
            n_lc     = sum(1 for r in recon_rows if r['flag'] == 'LATE CANCEL')
            n_np     = sum(1 for r in recon_rows if 'NO PUNCH' in r['flag'])
            n_over   = sum(1 for r in recon_rows if 'OVERBILLED' in r['flag'])
            n_under  = sum(1 for r in recon_rows if 'UNDERBILLED' in r['flag'])
            n_other  = total_items - n_match - n_lc - n_np - n_over - n_under

            c1, c2, c3, c4, c5, c6 = st.columns(6)
            c1.metric("Match",       n_match)
            c2.metric("No Punch",    n_np)
            c3.metric("Overbilled",  n_over)
            c4.metric("Underbilled", n_under)
            c5.metric("Other",       n_other)
            c6.metric("Late Cancel", n_lc)

            total_inv_hrs  = sum(r['inv_hrs'] for r in recon_rows if not r['lc'])
            total_emp_hrs  = sum(r['e_adj'] for r in recon_rows if r['e_adj'] is not None)
            punch_disc_hrs = round(sum(r['hrs_diff'] for r in recon_rows if r.get('hrs_diff') is not None), 2)
            missing_hrs    = round(sum(r['inv_hrs'] for r in recon_rows if r['match_type'] == 'no-match'), 2)

            hc1, hc2, hc3, hc4 = st.columns(4)
            hc1.metric("Total Invoice Hours", f"{total_inv_hrs:.2f}h")
            hc2.metric("Total Empion Hours",  f"{total_emp_hrs:.2f}h")
            disc_str = f"{punch_disc_hrs:+.2f}h"
            hc3.metric("Punch Discrepancy",   disc_str, delta=disc_str)
            hc4.metric("Missing Punch Hours", f"{missing_hrs:.2f}h")

            if name_notes:
                with st.expander(f"{len(name_notes)} name(s) matched with fuzzy/middle-name logic -- review"):
                    for note in name_notes:
                        st.write(f"* {note}")

            flag_rows = [r for r in recon_rows if r['flag'] not in ('MATCH', 'LATE CANCEL')]
            if flag_rows:
                st.subheader(f"{len(flag_rows)} Items to Review")
                import pandas as pd
                def fmt_dt(v):
                    return v.strftime("%-m/%-d %-I:%M %p") if v else "--"
                def conf_label(c):
                    if c is None: return "--"
                    if c >= 0.95: return f"OK {c:.0%}"
                    if c >= 0.80: return f"?? {c:.0%}"
                    return f"!! {c:.0%}"
                table_data = [{
                    "Date":       r['date'],
                    "Employee":   r['emp'],
                    "Role":       r['role'],
                    "Issue":      r['flag'],
                    "Confidence": conf_label(r.get('confidence')),
                    "Inv Hrs":     f"{r['inv_hrs']:.2f}",
                    "Emp Hrs":    f"{r['e_adj']:.2f}" if r['e_adj'] is not None else "--",
                    "Hrs Diff":   f"{r['hrs_diff']:+.2f}" if r['hrs_diff'] is not None else "--",
                    "Role":       r['role'],
                    "Issue":      r['flag'],
                    "Confidence": conf_label(r.get('confidence')),
                    "Inv Hrs":    f"{r['inv_hrs']:.2f}",
                    "Emp Hrs":    f"{r['e_adj']:.2f}" if r['e_adj'] is not None else "--",
                    "Hrs Diff":   f"{r['hrs_diff']:+.2f}" if r['hrs_diff'] is not None else "--",
                    "Inv In":     fmt_dt(r['start']),
                    "Inv Out":    fmt_dt(r['end']),
                    "Emp In":     fmt_dt(r['e_in']),
                    "Emp Out":    fmt_dt(r['e_out']),
                    "Notes":      r['notes'],
                } for r in flag_rows]
                st.dataframe(pd.DataFrame(table_data), use_container_width=True, hide_index=True)
            else:
                st.success("All shifts matched -- no issues found!")

            st.divider()
            st.subheader("Download Reports")
            inv_slug = meta["invoice"].replace("-", "_")
            fac_slug = re.sub(r"[^a-zA-Z0-9]", "_", meta["facility"])[:20].strip("_")

            d1, d2 = st.columns(2)
            with d1:
                hours_bytes = build_hours_excel(meta, shifts)
                st.download_button(
                    "Hours by Employee (Excel)",
                    data=hours_bytes,
                    file_name=f"{fac_slug}_{inv_slug}_Hours.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                flags_bytes = build_flags_excel(meta, recon_rows)
                st.download_button(
                    "Flags & Punch-for-Punch (Excel)",
                    data=flags_bytes,
                    file_name=f"{fac_slug}_{inv_slug}_Flags.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    else:
        st.info("Upload both files above to get started.")
